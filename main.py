from openpyxl import load_workbook

planilha = load_workbook('planilhas\Dados.xlsx')

# Definindo as páginas da planilha que serão utilizadas
dados_li = planilha["Dados LI"]
anuencias = planilha["Anuências"]

# Copiando Cabeçalho da pagina Anuencias
for cabecalho in range(2, 5):
    celula_anuencias = anuencias.cell(row=1, column=cabecalho)
    celula_dados_li = dados_li.cell(
        row=1, column=cabecalho+3, value=celula_anuencias.value)

print("---" * 15)
# Rodando código primeiro na página Dados LI
print("Dados LI:")
# Percorre todas as células mescladas na planilha
for mesclagem_dados_li in list(dados_li.merged_cells):
    # Remove a mesclagem
    dados_li.unmerge_cells(str(mesclagem_dados_li))
    # Após remover a mesclagem, algumas células das Colunas A, B e C ficarão vazias
print("Remoção de mesclagem bem sucedida.")
# Cria uma lista para armazenar quais linhas estão vazias na coluna A
linhas_para_apagar = []
for linhaDados in range(2, dados_li.max_row + 1):
    if dados_li.cell(row=linhaDados, column=1).value is None:
        linhas_para_apagar.append(linhaDados)
        
# Deleta as linhas que foram armazenadas (de trás para frente para evitar problemas de indexação)
for linhaDados in reversed(linhas_para_apagar):
    dados_li.delete_rows(linhaDados)
print("Linhas sem Número de Li apagadas")

# Rodando código na página Anuências
print("Anuências:")
# Percorre todas as células mescladas na planilha
for mesclagem_anuencias in list(anuencias.merged_cells):
    # Remove a mesclagem
    anuencias.unmerge_cells(str(mesclagem_anuencias))
    # Após remover a mesclagem, algumas células das Colunas A, B e C ficarão vazias
print("Remoção de mesclagem bem sucedida.")
# Cria uma lista para armazenar quais linhas estão vazias na coluna A
linhas_para_excluir = []
for linhaAnuencias in range(2, anuencias.max_row + 1):
    if anuencias.cell(row=linhaAnuencias, column=1).value is None:
        linhas_para_excluir.append(linhaAnuencias)

# Deleta as linhas que foram armazenadas (de trás para frente para evitar problemas de indexação)
for linhaAnuencias in reversed(linhas_para_excluir):
    anuencias.delete_rows(linhaAnuencias)
print("Linhas sem Número de Li Apagadas")

# Dicionario para mapear as LIs em cada linha
id_dados_li = {}
id_anuencias = {}

# Mapeando LIs na Planilha Dados LI
for linhaDados in range(2, dados_li.max_row + 1):
    li1 = dados_li.cell(row=linhaDados, column=1).value
    if li1 not in id_dados_li:
        id_dados_li[li1] = []
    id_dados_li[li1].append(linhaDados)

# Mapeando LIs na Planilha Anuências
for linhaAnuencias in range(2, anuencias.max_row + 1):
    li2 = anuencias.cell(row=linhaAnuencias, column=1).value
    if li2 not in id_anuencias:
        id_anuencias[li2] = []
    id_anuencias[li2].append(linhaAnuencias)

# Copiando os dados das Anuências para Dados LI onde os IDs são iguais
for li_iguais in set(id_dados_li.keys()).intersection(id_anuencias.keys()):
    linhasDados = id_dados_li[li_iguais]
    linhasAnuencias = id_anuencias[li_iguais]
    
    for linhaDados, linhaAnuencias in zip(linhasDados, linhasAnuencias):
        # Copiar os dados das colunas B, C e D de Anuências para E, F e G de Dados LI
        for coluna in range(2, 5):
            valor_anuencias = anuencias.cell(row=linhaAnuencias, column=coluna).value
            dados_li.cell(row=linhaDados, column=coluna + 3).value = valor_anuencias

planilha.save('planilhas\DadosFinal.xlsx')
print("Finalizado")

